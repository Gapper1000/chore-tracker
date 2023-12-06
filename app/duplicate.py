from openpyxl import load_workbook
filePath = "/Users/gmp/Documents/GitHub/chore-tracker/Chore Tracker.xlsx"
workbook = load_workbook(filePath)
source_sheet_name = workbook.sheetnames[0]
source_sheet = workbook[source_sheet_name]
duplicated_sheet_name = f"Copy of {source_sheet_name}"
duplicated_sheet = workbook.copy_worksheet(source_sheet)
duplicated_sheet.title = duplicated_sheet_name
workbook.save(filePath)
workbook.close()