import pandas as pd
import os
from openpyxl import load_workbook

filePath = "/Users/gmp/Documents/GitHub/chore-tracker/Chore Tracker.xlsx"
selected_columns = ['Assignee', 'Task', 'Due Date', 'Status']


def readFile(filePath):
        chore_df = pd.read_excel(filePath, usecols=selected_columns)
        return chore_df

def editStatus(chore_df, task_name=str, new_status=bool):
        index = columnIndex(chore_df, 'Task', task_name)
        if index != None:
            chore_df.loc[index, 'Status'] = new_status
        return chore_df

def columnIndex(chore_df, column= str, cell= str):
        try:
            index = chore_df.index[chore_df[column] == cell].tolist()[0]
            return index
        except IndexError:
            print(f"{column} {cell} not found please try again")
            return None
        
def saveToFile(chore_df, filePath):
        chore_df.to_excel(filePath, index=False)
        return

def choreProgress(chore_df, status_column='Status'):
    try:
        if status_column not in chore_df.columns:
            raise ValueError(f"Column '{status_column}' not found in the Excel sheet.")

            # Convert the status column to numeric (assuming it contains numeric values)
        chore_df[status_column] = pd.to_numeric(chore_df[status_column], errors='coerce')

            # Sum up the progress values
        progress_sum = chore_df[status_column].sum()

        return progress_sum

    except Exception as e:
        print(f"Error: {e}")
        return None
    
def duplicateSheet(filePath): #doesnt work here but works in duplicate.py
        try:
            workbook = load_workbook(filePath)
            source_sheet_name = workbook.sheetnames[0]
            source_sheet = workbook[source_sheet_name]
            duplicated_sheet_name = f"Copy of {source_sheet_name}"
            duplicated_sheet = workbook.copy_worksheet(source_sheet)
            duplicated_sheet.title = duplicated_sheet_name
            workbook.save(filePath)
            workbook.close()
            return

        except Exception as e:
            print(f"Error duplicating sheet: {e}")
            return

def update_workbook(file_path, checked_tasks_ids):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=False):
        task_id = row[0].value
        # Set the value to True if task_id is in checked_tasks_ids, else False
        row[3].value = str(task_id) in checked_tasks_ids

    workbook.save(filename=file_path)
